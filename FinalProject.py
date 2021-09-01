from os import name
import tkinter
import pyodbc
import openpyxl
from tkinter import messagebox
from tkinter import *

class Customer:
    def __init__(self, name, budget):
        self.name = name
        self.budget = budget
        self.finalBill = {}

    def BuyTicket(self, cinema):
        movieIndex = 0
        print("Please choose the movie you want to watch: ")
        cinema.printMovies()
        movieIndex = int(input()) - 1
        numOfTickets = int(input("How many tickets do you want to buy?"))
        self.budget -= (cinema.avilableMovies[movieIndex].price * numOfTickets)
        self.AddToBill(cinema.avilableMovies[movieIndex].movieName ,cinema.avilableMovies[movieIndex].price, numOfTickets)
        f = open('MovieTicket.txt','a')
        f.write(self.name+ ' purchase ' + str(numOfTickets) + ' tickets to the movie ' + cinema.avilableMovies[movieIndex].movieName + ' for ' + str(cinema.avilableMovies[movieIndex].price * numOfTickets) + '$\n' )
        f.close()
        print("You bought " + str(numOfTickets) + " tickets to " + cinema.avilableMovies[movieIndex].movieName + ", money left: " + str(self.budget) + "$")
        
    def BuySnack(self, buffet):
        snackIndex = 0
        print("Buffet menu: ")
        buffet.printBuffet()
        snackIndex = int(input()) - 1
        numOfSnacks = int(input("How many products do you want to buy?"))
        self.budget -= (buffet.avilableProducts[snackIndex].snackPrice * numOfSnacks)
        print("You bought " + buffet.avilableProducts[snackIndex].snackName + ", money left: " + str(self.budget) + "$")
        self.AddToBill(buffet.avilableProducts[snackIndex].snackName,buffet.avilableProducts[snackIndex].snackPrice,numOfSnacks)

    def AddToBill(self, product, productPrice, productAmount):
        for i in self.finalBill:
            if product in i:
                self.finalBill[product] = (productPrice, self.finalBill[product][1] + productAmount)
                return 
        self.finalBill[product] = (productPrice, productAmount)

    def ShowFinalBill(self):
	    sum = 0
	    for product, details in self.finalBill.items():
		    sum += details[1] * details[0]
		    print(product + " ,quantity: " + str(details[1]) + " ,price: " + str(details[1] * details[0]) + "$")
	    print("Total cost: " + str(sum)+"$")

class Movie:
    def __init__(self, code, movieName, category, ScreeningDate, price):
        self.code = code
        self.movieName = movieName
        self.category = category
        self.ScreeningDate = ScreeningDate
        self.price = price
    
class Snack:
    def __init__(self,snackName, snackPrice):
        self.snackName = snackName
        self.snackPrice = snackPrice

class Buffet:
    def __init__(self):
        self.avilableProducts = [Snack("Coca cola",15),
        Snack("Water",5),
        Snack("Popcorn",20),
        Snack("Nachos",15),
        Snack("Chocolate",10),
        Snack("Bamba",10)]

    def printBuffet(self):
        j = 1
        print("********************")
        print("Buffet Menu:")
        for i in self.avilableProducts :
            print(str(j) + ". " + i.snackName +", Price: " + str(i.snackPrice) + "$")
            j +=1   
            
class Cinema:
    def __init__(self):
        self.avilableMovies = [Movie(536, "Fast & Furious", "Action", "24/06/2021", 30), 
        Movie(245, "Lion King", "Disney", "24/06/2021",20), 
        Movie(875, "Aladdin", "Disney", "25/07/2021", 20), 
        Movie(444, "Taken", "Action", "30/06/2021", 30), 
        Movie(333, "Neighbors", "Comedy", "26/06/2021", 25), 
        Movie(555, "We're the Millers", "Comedy", "01/07/2021", 25), 
        Movie(809, "Beauty and the Beast", "Disney", "29/06/2021", 20), 
        Movie(666, "American Sniper", "Action", "25/06/2021", 30), 
        Movie(213, "Django Unchained", "Action", "01/07/2021", 30), 
        Movie(617, "The Little Mermaid", "Disney", "03/07/2021", 20), 
        Movie(321, "The Hangover", "Comedy", "05/07/2021", 25), 
        Movie(893, "American Pie", "Comedy", "10/07/2021", 25), 
        Movie(445, "Mulan", "Disney", "07/07/2021", 20)]

    def printMovies(self):
        print("Avilable movies in our Cinema:")
        j = 1
        for i in self.avilableMovies :
            print(str(j) + ". Name: " + i.movieName + ", Category: " + i.category + ", Date: " + i.ScreeningDate + " , Price: " + str(i.price) + "$")
            j +=1

def menu():
    def ExcNoneQuery(sql_command, values):
        server= 'LAPTOP-K8MAK0VU\SQLEXPRESS'
        database = 'Movies'
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; \
                                SERVER=' + server +'; \
                                DATABASE='+ database +'; \
                                Trusted_connection=yes;')
        crsr = cnxn.cursor()
        crsr.execute(sql_command, values)
        crsr.commit()
        crsr.close()
        cnxn.close()
    
    def ExcQuery(sql_command):
        server= 'LAPTOP-K8MAK0VU\SQLEXPRESS'
        database = 'Movies'
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; \
                            SERVER=' + server +'; \
                            DATABASE='+ database +'; \
                            Trusted_connection=yes;')
        crsr = cnxn.cursor()
        crsr.execute(sql_command)                     
        for row in crsr:
            print(row)
        crsr.close()
        cnxn.close()       

    window = Tk()
    window.title("Nitzan & Eliran")
    bg1 = PhotoImage(file=r'''C:\Users\Nitzan Gabay\OneDrive - Ruppin Academic Center\Desktop\שנה ב'\סמסטר ב\פייטון\מבחן מסכם בפייטון\Popcorn.png''')
    bg1_lable = Label(window, image= bg1)
    bg1_lable.place(x=0, y=0, relwidth=1, relheight=1)
    lableName = Label(window, text="Name:", font=("Sitka Small", 30), bg="black", fg="white" )
    lableName.grid(column=8, row=1)
    userName = Entry(window,font=("Sitka Small", 20),width=5)
    userName.grid(column=8, row=2)
    lableBudget = Label(window, text="Budget:", font=("Sitka Small", 30), bg="black", fg="white")
    lableBudget.grid(column=8, row=3)
    userBudget = Entry(window, font=("Sitka Small", 20),width=5)
    userBudget.grid(column=8, row=4)
    lable = Label(window, text="", font=("David bold", 10))
    lable.grid(column=0, row=5)
    lable = Label(window, text="", font=("David bold", 10))
    lable.grid(column=0, row=6)
    def clickToStart():
        messagebox.showinfo("Setup", "Hello " + userName.get() + "!\nWelcome to our Cinema!")
        global name1
        global budget1
        name1 = userName.get()
        budget2 = userBudget.get()
        budget1 = int(budget2)
        window.destroy() 
    bt = Button(window, text = "Submit",font=("Sitka Small", 15), bg='black', fg='white', width=10, command = clickToStart)
    bt.grid(column=5, row=8)
    window.geometry('480x330')
    window.mainloop()

    customers = Customer(name1, budget1)
    cinema = Cinema()
    buffet = Buffet()

    def clickBtn1():
        cinema.printMovies()

    def clickBtn2():
        customers.BuyTicket(cinema)

    def clickBtn3():
        buffet.printBuffet()

    def clickBtn4():
        customers.BuySnack(buffet)

    def clickBtn5():
        print("*****************\n")
        print("Your final bill: ")
        customers.ShowFinalBill()

    def clickBtn6():
        sql_command = 'Insert into Customers(Name,Budget) VALUES(?,?)'
        values = (name1,budget1)
        ExcNoneQuery(sql_command, values)
        print("The customer " + name1 + " saved to DB")   

    def clickBtn7():
        sql_command = 'Select * from Customers Order By Id'
        ExcQuery(sql_command)       

    def clickBtn8():
        wb = openpyxl.load_workbook(filename = './Movies.xlsx')
        sheet = wb['Movies']
        i = 1
        for item in cinema.avilableMovies:
            sheet.cell(row = i, column = 1).value = str(item.movieName)
            sheet.cell(row = i, column = 2).value = str(item.category)
            sheet.cell(row = i, column = 3).value = str(item.ScreeningDate)
            i += 1
        wb.save('Movies.xlsx')    
        print(str(i-1) + " Movies added to Excel file.")

    def clickBtn9():
        print("Bye Bye..")
        newMenu.destroy()

    newMenu = Tk()
    newMenu.title("Nitzan & Eliran")
    bg = PhotoImage(file=r'''C:\Users\Nitzan Gabay\OneDrive - Ruppin Academic Center\Desktop\שנה ב'\סמסטר ב\פייטון\מבחן מסכם בפייטון\Cinema1.png''')
    bg_lable = Label(newMenu, image= bg)
    bg_lable.place(x=0, y=0, relwidth=1, relheight=1)
    headline = Label(newMenu, text = " Ruppin Cinema ", font=("Castellar", 49), fg="White", bg="black")
    headline.pack(pady=40)
    my_frame = Frame(newMenu)
    my_frame.pack(pady=0)
    bt1 = Button(my_frame, text = "See Avilable Movies",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn1)
    bt1.grid(row = 1, column = 0)
    bt2 = Button(my_frame, text = "Buy a Ticket",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn2)
    bt2.grid(row = 1, column = 1)
    bt3 = Button(my_frame, text = "Show Buffet Menu",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn3)
    bt3.grid(row = 2, column = 0)
    bt4 = Button(my_frame, text = "Buy a Snack",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn4)
    bt4.grid(row = 2, column = 1)
    bt5 = Button(my_frame, text = "Show Finall Bill",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn5)
    bt5.grid(row = 3, column = 0)
    bt6 = Button(my_frame, text = "Save Customer To DB",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn6)
    bt6.grid(row = 3, column = 1)
    bt7 = Button(my_frame, text = "Print Customers From DB",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn7)
    bt7.grid(row = 4, column = 0)
    bt8 = Button(my_frame, text = "Save Movies To Excel",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn8)
    bt8.grid(row = 4, column = 1)
    bt9 = Button(newMenu, text = "Exit",font=("Sitka Small", 16), bg='black', fg='white', height=1, width=20, command = clickBtn9)
    bt9.pack(padx=0, pady= 20)

    newMenu.geometry('605x580')
    newMenu.mainloop()
menu()

#Pandas

import pandas as pd
import numpy as np
import csv
import seaborn as sns
import matplotlib.pyplot as plt
#1
df = pd.read_csv("wishSalesSummer2020.csv")
display(df)
#2
df.shape
#3
df.info()
#4
del df['product_picture']
#פקודה מסייעת לפקודה 10
df['product_color'] = df['product_color'].str.capitalize()
#5
newTable = df.filter(['title_orig', 'price', 'retail_price', 'units_sold'], axis = 1)
display(newTable)
#6
newTable["revenue"] = newTable['retail_price']*newTable['units_sold']
display(newTable)
#7
newTable["profit"] = ((newTable['retail_price'] - newTable['price']) * newTable['units_sold'])
display(newTable)
#8
newTable.sort_values(by = 'profit', ascending = False)
#9
newTable2 = newTable[newTable['profit']<0]
lost = newTable2['profit'].sum()
print("The company lost for aug 2020 is: " + str(lost))
#10
newTable3 = df.filter(['retail_price', 'units_sold','product_color'], axis = 1)
newTable4 = newTable3.groupby(['product_color'])[['units_sold']].sum()
newTable4.sort_values(by = 'units_sold', ascending = False)
#11
newTable['merchant_name'] = df['merchant_name']
a = newTable.groupby(['merchant_name'])[['profit']].sum()
a.sort_values(by = 'profit', ascending = False)
meanProfit = round(a['profit'].mean())
print("The mean profit of every merchant is : " + str(meanProfit))
#12
df.groupby(['product_size'])[['units_sold']].sum()
df.groupby(['product_size'])[['units_sold']].sum().max()
#13
newTable["difference"] = (newTable['retail_price'] - newTable['price'])
newTable.sort_values(by = 'difference', ascending = False)
#14
merchants = df.filter(['merchant_name', 'merchant_rating_count', 'merchant_rating'], axis = 1)
display(merchants)
#15
ratingTable = merchants[ merchants['merchant_rating_count'] > 50]
ratingTable.sort_values(by = 'merchant_rating', ascending = False)
#16
merchantMean = ratingTable['merchant_rating'].mean()
print("The mean merchant_rating of every merchant is : " + str(merchantMean))
#17
newTable[['units_sold','retail_price']].agg(['min','max','mean','std']).round()
#18
KamaOd = df.filter(['title_orig', 'countries_shipped_to','product_color','product_size'], axis = 1)
display(KamaOd[KamaOd.countries_shipped_to == KamaOd.countries_shipped_to.max()])

#seaborn
cols = pd.read_csv("wishSalesSummer2020.csv").columns
wish = pd.read_csv("wishSalesSummer2020.csv")
#19
sns.relplot(data=wish, x="origin_country", y="units_sold").set(title='Units sales per country')
#20
sns.relplot(x="origin_country",y="product_size",data=wish).set(title='Size per coutry')
